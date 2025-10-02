#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Rolling Factor Regressions - (Returns from Excel, 4F built from 3F+Momentum CSVs, FF5 from CSV)
- Returns (DECIMALS) from Excel sheet
- Carhart 4F built at runtime by merging F-F 3-Factor CSV with Momentum CSV (UMD)
- FF5 from F-F 5-Factors 2x3 CSV
- Outputs written to Excel
"""

import pandas as pd, numpy as np, re
import statsmodels.api as sm
from pathlib import Path

# ================= USER CONFIG =================
# ---- Returns (Excel) ----
EXCEL_RETURNS_PATH = "/Users/hugofayolle/Downloads/SP500_Stocks_and_Prices_per_month/SP500_prices_matrix.xlsx"
RETURNS_SHEET      = "Backtest_Returns"   # robust resolver will try variants including "Backest returns"

# ---- Factors (CSVs) ----
FACTOR_CSV_PATH_3F = "/Users/hugofayolle/Downloads/F-F_Research_Data_Factors.csv"       # Date, Mkt-RF, SMB, HML, RF
MOMENTUM_CSV_PATH  = "/Users/hugofayolle/Downloads/F-F_Momentum_Factor.csv"             # Date, Mom (UMD)
FACTOR_CSV_PATH_5F = "/Users/hugofayolle/Downloads/F-F_Research_Data_5_Factors_2x3.csv" # Date, Mkt-RF, SMB, HML, RMW, CMA, RF

# ---- Output workbook (can be same as returns Excel) ----
OUTPUT_EXCEL_PATH  = EXCEL_RETURNS_PATH

# ---- Output sheet names ----
OUT_SUMMARY_4F = "Factor_Regs_Summary_4F"
OUT_ROLLING_4F = "Factor_Regs_Rolling_4F"
OUT_BETAS_4F   = "Factor_Regs_Betas_4F"
OUT_DIFF_4F    = "Factor_Diff_4F"

OUT_SUMMARY_5F = "Factor_Regs_Summary_FF5"
OUT_ROLLING_5F = "Factor_Regs_Rolling_FF5"
OUT_BETAS_5F   = "Factor_Regs_Betas_FF5"
OUT_DIFF_5F    = "Factor_Diff_FF5"

# ---- Window & stats ----
START_MONTH = "2015-01"
END_MONTH   = "2025-08"
ROLL_WIN    = 36   # months
HAC_LAGS    = 3    # Newey–West lags
# ===========================================================


# ---------------------------- Utilities ----------------------------
def _ensure_month_col(df: pd.DataFrame) -> pd.DataFrame:
    """Create/normalize 'month' to YYYY-MM without noisy dateutil warnings."""
    df = df.copy()

    def norm_month_series(s: pd.Series) -> pd.Series:
        s = s.astype(str).str.strip()
        s = s.str.replace(r"[^\d-]", "", regex=True)
        s = s.apply(lambda x: f"{x[:4]}-{x[4:6]}" if re.fullmatch(r"\d{6}", x or "") else x)
        return s

    if "month" in df.columns:
        df["month"] = norm_month_series(df["month"])
        return df

    # try index first
    if df.index.name:
        idx = df.index.astype(str)
        if pd.Series(idx).str.fullmatch(r"\d{6}").any():
            df["month"] = pd.Series(idx).str.replace(r"[^\d]", "", regex=True).str.replace(
                r"^(\d{4})(\d{2})$", r"\1-\2", regex=True
            )
            return df
        if pd.Series(idx).str.fullmatch(r"\d{4}-\d{2}").any():
            df["month"] = idx
            return df

    # fallback: first column
    first = df.columns[0]
    df["month"] = norm_month_series(df[first])
    return df


def _clean_ff_cols(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty: return df
    new_cols = []
    for c in df.columns:
        cc = str(c).strip()
        if cc.lower().startswith("mom"):
            cc = "UMD"
        cl = cc.lower().replace(" ", "").replace("-", "").replace("_", "")
        if cl in {"mktrf","mktminusrf","marketrf"}:
            cc = "Mkt-RF"
        elif cl == "smb": cc = "SMB"
        elif cl == "hml": cc = "HML"
        elif cl == "rmw": cc = "RMW"
        elif cl == "cma": cc = "CMA"
        elif cl == "rf":  cc = "RF"
        new_cols.append(cc)
    out = df.copy()
    out.columns = new_cols
    return out


def _coerce_numeric(df: pd.DataFrame, cols) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def mask_window(idx, start, end):
    idx = pd.Index(idx.astype(str))
    if start is None and end is None:
        return pd.Series(True, index=idx)
    lo = start if start is not None else idx.min()
    hi = end   if end   is not None else idx.max()
    if lo > hi: lo, hi = hi, lo
    return (idx >= lo) & (idx <= hi)


def _sniff_data_start(text: str):
    for i, line in enumerate(text.splitlines()):
        if re.match(r"^\s*\d{6}\s*[,\t]", line):   # 201001,<...>
            return i
        if re.match(r"^\s*\d{4}-\d{2}\s*[,\t]", line):  # 2010-01,<...>
            return i
    return None


def _read_ff_csv_robust(path: str) -> pd.DataFrame:
    raw = open(path, "r", encoding="utf-8", errors="replace").read()
    start = _sniff_data_start(raw)
    if start is None:
        # try normal read; user may have a clean CSV already
        df = pd.read_csv(path)
    else:
        lines = raw.splitlines()
        header_idx = start - 1 if start > 0 else None
        if header_idx is not None and ("," in lines[header_idx] or "\t" in lines[header_idx]):
            df = pd.read_csv(path, skiprows=header_idx, engine="python")
        else:
            df = pd.read_csv(path, skiprows=start, header=None, engine="python")
    # ensure month col, clean names
    df = _ensure_month_col(df)
    df = _clean_ff_cols(df)
    return df


# ---------------------------- Loaders ----------------------------
def _resolve_returns_sheet(path: str, desired: str) -> str:
    xls = pd.ExcelFile(path, engine="openpyxl")
    sheets = xls.sheet_names
    # exact
    if desired in sheets: return desired
    # variants
    candidates = [desired, "Backtest returns","Backtest_Returns","Backest returns","Backest Returns","Backest_returns"]
    for c in candidates:
        if c in sheets: return c
    # fuzzy
    def norm(s): return str(s).lower().replace(" ", "").replace("_", "")
    nd = norm(desired)
    for s in sheets:
        if norm(s) == nd: return s
    raise ValueError(f"Returns sheet not found. Looked for '{desired}' & variants. Found: {sheets}")


def load_returns_from_excel() -> pd.DataFrame:
    sheet = _resolve_returns_sheet(EXCEL_RETURNS_PATH, RETURNS_SHEET)
    r = pd.read_excel(EXCEL_RETURNS_PATH, sheet_name=sheet, engine="openpyxl")
    # robust rename
    low = {c.lower(): c for c in r.columns}
    ren = {}
    if "month" not in r.columns:
        k = low.get("month") or low.get("date")
        if k: ren[k] = "month"
    if "base_ret" not in r.columns:
        k = low.get("base_ret") or low.get("base") or low.get("base_return")
        if k: ren[k] = "base_ret"
    if "qual_ret" not in r.columns:
        k = low.get("qual_ret") or low.get("qual") or low.get("qual_return")
        if k: ren[k] = "qual_ret"
    if ren: r = r.rename(columns=ren)
    need = {"month","base_ret","qual_ret"}
    if not need.issubset(r.columns):
        raise ValueError(f"Returns sheet must include {need}, found {list(r.columns)}")
    r = _ensure_month_col(r).set_index("month").sort_index()
    r = r.loc[mask_window(r.index, START_MONTH, END_MONTH)]
    # numeric coercion (DECIMALS)
    r["base_ret"] = pd.to_numeric(r["base_ret"], errors="coerce")
    r["qual_ret"] = pd.to_numeric(r["qual_ret"], errors="coerce")
    return r[["base_ret","qual_ret"]]


def load_4f_from_3f_plus_mom() -> pd.DataFrame:
    """Build Carhart 4F by merging 3F CSV with Momentum CSV on 'month'."""
    f3 = _read_ff_csv_robust(FACTOR_CSV_PATH_3F)
    mom = _read_ff_csv_robust(MOMENTUM_CSV_PATH)
    # Restrict to needed cols
    cand3 = [c for c in ["month","Mkt-RF","SMB","HML","RF"] if c in f3.columns]
    f3 = f3[cand3].copy()
    # Momentum: ensure 'UMD' exists
    if "UMD" not in mom.columns:
        # find a column that starts with mom
        mom_col = None
        for c in mom.columns:
            if c != "month" and str(c).lower().startswith("mom"):
                mom_col = c
                break
        if mom_col is None and len(mom.columns) >= 2:
            mom_col = mom.columns[1]
        mom = mom.rename(columns={mom_col: "UMD"})
    mom = mom[["month","UMD"]].copy()
    # Normalize month formats
    f3 = _ensure_month_col(f3)
    mom = _ensure_month_col(mom)
    # Merge and window
    four = pd.merge(f3, mom, on="month", how="inner").sort_values("month")
    four = four.set_index("month").sort_index()
    four = four.loc[mask_window(four.index, START_MONTH, END_MONTH)]
    # numeric coercion (PERCENT)
    four = _coerce_numeric(four, ["Mkt-RF","SMB","HML","UMD","RF"])
    return four


def load_ff5_csv() -> pd.DataFrame:
    f5 = _read_ff_csv_robust(FACTOR_CSV_PATH_5F)
    need = {"month","Mkt-RF","SMB","HML","RMW","CMA","RF"}
    if not need.issubset(f5.columns) and f5.shape[1] >= 7:
        f5 = f5.iloc[:, :7]
        f5.columns = ["month","Mkt-RF","SMB","HML","RMW","CMA","RF"]
    f5 = f5.set_index("month").sort_index()
    f5 = f5.loc[mask_window(f5.index, START_MONTH, END_MONTH)]
    # numeric coercion (PERCENT)
    f5 = _coerce_numeric(f5, ["Mkt-RF","SMB","HML","RMW","CMA","RF"])
    return f5


# ---------------------------- Stats helpers ----------------------------
def fit_ols_hac(y_pct, X_pct, hac_lags=HAC_LAGS):
    """Return (alpha_pm, t_alpha, betas_dict, R2). Inputs in PERCENT."""
    df = pd.concat([y_pct, X_pct], axis=1).dropna()
    if df.empty:
        return np.nan, np.nan, {}, np.nan
    y = df.iloc[:, 0]
    X = sm.add_constant(df.iloc[:, 1:])
    res = sm.OLS(y, X).fit(cov_type="HAC", cov_kwds={"maxlags": hac_lags})
    alpha = res.params.get("const", np.nan)
    tstat = res.tvalues.get("const", np.nan)
    betas = {k: res.params[k] for k in res.params.index if k != "const"}
    return alpha, tstat, betas, res.rsquared


def rolling_alpha(y_pct, X_pct, win=ROLL_WIN, hac_lags=HAC_LAGS):
    """Rolling alpha & t (HAC)."""
    df = pd.concat([y_pct, X_pct], axis=1).dropna()
    if df.shape[0] < win:
        return pd.DataFrame(columns=["alpha_pm","t_alpha"])
    out = []
    for i in range(win, len(df)+1):
        yi = df.iloc[i-win:i, 0]
        Xi = df.iloc[i-win:i, 1:]
        a, t, _, _ = fit_ols_hac(yi, Xi, hac_lags)
        out.append((df.index[i-1], a, t))
    return pd.DataFrame(out, columns=["month","alpha_pm","t_alpha"]).set_index("month")


def ann_from_monthly_alpha(alpha_pm):
    if pd.isna(alpha_pm): return np.nan
    return (1 + alpha_pm/100.0)**12 - 1


# ---------------------------- Model runner ----------------------------
def run_model(tag: str, factors_pct: pd.DataFrame, r_dec: pd.DataFrame):
    """
    tag: '4F' or 'FF5'
    factors_pct: factors in PERCENT incl RF
    r_dec: returns in DECIMALS (base_ret, qual_ret)
    """
    idx = r_dec.index.intersection(factors_pct.index)
    if len(idx) < 24:
        raise ValueError(f"Not enough overlap between returns and {tag} factors.")
    f = factors_pct.loc[idx].copy()

    base_pct = (r_dec.loc[idx, "base_ret"].astype(float) * 100.0)
    qual_pct = (r_dec.loc[idx, "qual_ret"].astype(float) * 100.0)

    if tag == "4F":
        X = f[["Mkt-RF","SMB","HML","UMD"]]
        out_sum, out_roll, out_betas, out_diff = (OUT_SUMMARY_4F, OUT_ROLLING_4F, OUT_BETAS_4F, OUT_DIFF_4F)
        model_name = "Carhart_4F"
    else:
        X = f[["Mkt-RF","SMB","HML","RMW","CMA"]]
        out_sum, out_roll, out_betas, out_diff = (OUT_SUMMARY_5F, OUT_ROLLING_5F, OUT_BETAS_5F, OUT_DIFF_5F)
        model_name = "FF5"

    xs_base = base_pct - f["RF"]   # percent excess
    xs_qual = qual_pct - f["RF"]

    # Rolling alphas
    roll_b = rolling_alpha(xs_base.loc[X.index], X)
    roll_q = rolling_alpha(xs_qual.loc[X.index], X)
    rolling = pd.DataFrame({
        "month": roll_b.index,
        "alpha_base_pm": roll_b["alpha_pm"].values,
        "t_base": roll_b["t_alpha"].values,
        "alpha_qual_pm": roll_q["alpha_pm"].reindex(roll_b.index).values,
        "t_qual": roll_q["t_alpha"].reindex(roll_b.index).values,
    })

    # Eras
    eras = [
        ("Full",       X.index.min(), X.index.max()),
        ("2010-2014",  "2010-01",     "2014-12"),
        ("2015-2025",  "2015-01",     X.index.max()),
    ]
    def _mask(ix, s, e):
        ii = pd.Index(ix.astype(str))
        if s is None: s = ii.min()
        if e is None: e = ii.max()
        return (ii >= s) & (ii <= e)

    sum_rows, beta_rows = [], []
    for label, s, e in eras:
        sel = X.index[_mask(X.index, s, e)]
        if len(sel) < 24:
            continue
        yb = xs_base.loc[sel]; yq = xs_qual.loc[sel]; Xe = X.loc[sel]
        a_b, t_b, b_b, r2_b = fit_ols_hac(yb, Xe)
        a_q, t_q, b_q, r2_q = fit_ols_hac(yq, Xe)
        sum_rows.append({
            "Model": model_name, "Era": label, "From": sel.min(), "To": sel.max(),
            "Alpha_Base_pm_%": a_b, "t_Base": t_b, "Alpha_Base_ann": ann_from_monthly_alpha(a_b),
            "Alpha_Qual_pm_%": a_q, "t_Qual": t_q, "Alpha_Qual_ann": ann_from_monthly_alpha(a_q),
            "R2_Base": r2_b, "R2_Qual": r2_q,
            "Alpha_Diff_pm_%": (a_q - a_b)
        })
        beta_rows += [
            {"Model": model_name, "Era": label, "Portfolio": "Base", **b_b},
            {"Model": model_name, "Era": label, "Portfolio": "Qual", **b_q},
        ]

    summary = pd.DataFrame(sum_rows)
    betas   = pd.DataFrame(beta_rows)

    # Diff model (Qual - Base), full-sample + rolling
    diff = (qual_pct - base_pct)
    a_d, t_d, b_d, r2_d = fit_ols_hac(diff.loc[X.index], X)
    diff_full = pd.DataFrame([{
        "Model": model_name, "Era": "Full",
        "Alpha_Diff_pm_%": a_d, "t_Diff": t_d,
        "Alpha_Diff_ann": ann_from_monthly_alpha(a_d),
        "R2_Diff": r2_d
    }])

    roll_d = rolling_alpha(diff.loc[X.index], X)
    if not roll_d.empty:
        roll_d["alpha_ann"] = (1 + roll_d["alpha_pm"]/100.0)**12 - 1
        roll_d = roll_d.reset_index().rename(columns={"index":"month"})

    return (summary, rolling, betas, diff_full, roll_d, out_sum, out_roll, out_betas, out_diff)


# ---------------------------- Main ----------------------------
def main():
    # Load inputs
    r   = load_returns_from_excel()       # decimals
    f4  = load_4f_from_3f_plus_mom()      # percent (built from 3F+Mom)
    f5  = load_ff5_csv()                  # percent

    # Run models
    res4 = run_model("4F",  f4, r)
    res5 = run_model("FF5", f5, r)

    # Purge prior result sheets to avoid old+new coexistence
    try:
        from openpyxl import load_workbook
        wb = load_workbook(OUTPUT_EXCEL_PATH)
        targets = [
            OUT_SUMMARY_4F, OUT_ROLLING_4F, OUT_BETAS_4F, OUT_DIFF_4F, OUT_DIFF_4F + "_Rolling",
            OUT_SUMMARY_5F, OUT_ROLLING_5F, OUT_BETAS_5F, OUT_DIFF_5F, OUT_DIFF_5F + "_Rolling",
        ]
        for name in list(wb.sheetnames):
            if name in targets:
                wb.remove(wb[name])
        wb.save(OUTPUT_EXCEL_PATH)
    except FileNotFoundError:
        pass

    # Write results (replace sheets if they exist)
    with pd.ExcelWriter(OUTPUT_EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
        # 4F
        res4[0].to_excel(w, index=False, sheet_name=res4[5])
        res4[1].to_excel(w, index=False, sheet_name=res4[6])
        res4[2].to_excel(w, index=False, sheet_name=res4[7])
        res4[3].to_excel(w, index=False, sheet_name=res4[8])
        if not res4[4].empty:
            res4[4].to_excel(w, index=False, sheet_name=res4[8] + "_Rolling")

        # FF5
        res5[0].to_excel(w, index=False, sheet_name=res5[5])
        res5[1].to_excel(w, index=False, sheet_name=res5[6])
        res5[2].to_excel(w, index=False, sheet_name=res5[7])
        res5[3].to_excel(w, index=False, sheet_name=res5[8])
        if not res5[4].empty:
            res5[4].to_excel(w, index=False, sheet_name=res5[8] + "_Rolling")

    print("Done. Wrote sheets:",
          OUT_SUMMARY_4F, OUT_ROLLING_4F, OUT_BETAS_4F, OUT_DIFF_4F, OUT_DIFF_4F + "_Rolling", "and",
          OUT_SUMMARY_5F, OUT_ROLLING_5F, OUT_BETAS_5F, OUT_DIFF_5F, OUT_DIFF_5F + "_Rolling")


if __name__ == "__main__":
    main()
